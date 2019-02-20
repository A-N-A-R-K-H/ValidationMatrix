import pandas as pd
import numpy as np
import os
import glob
import math
import openpyxl
import datetime
import getpass


# Get row offset by given PCB
def pcb_row_offset(pcb):

    row = {
        "T581806175": 0,  # PCB8
        "T581806177": 1,  # PCB10
        "T581806168": 2,  # PCB11
        "T581806170": 3,  # PCB12
    }
    return row.get(pcb, None)


# Get row by given Matrix ID
def matrix_id_row(id):

    row = {
        "M1": 5,
        "M2": 10,
        "M3": 15,
        "M4": 20,
        "M5": 25,
        "M6": 30,
        "M7": 35,
        "M8": 40,
        "M9": 45,
        "M10": 50,
        "M11": 55,
        "M12": 60,
        "M13": 65,
        "M14": 70,
        "M15": 75,
        "M16": 80,
        "M17": 85,
        "M18": 90,
    }
    return row.get(id, None)


# Get a lits of positions by given PCB
def get_fe_list(pcb):

    fe_list = {
        # PCB8
        "T581806175": {"Low": {"TX": [8, 8], "RX": [8, 8]}, "Mid": {"TX": [12, 12], "RX": [12, 12]}, "Hig": {"TX": [1, 1], "RX": [1, 1]}},
        # PCB10
        "T581806177": {"Low": {"TX": [7, 5], "RX": [7, 5]}, "Mid": {"TX": [15, 13], "RX": [4, 11]}, "Hig": {"TX": [1, 0], "RX": [1, 0]}},
        # PCB11
        "T581806168": {"Low": {"TX": [0, 5], "RX": [0, 5]}, "Mid": {"TX": [12, 14], "RX": [5, 14]}, "Hig": {"TX": [15, 0], "RX": [15, 15]}},
        # Dummy
        "T581806170": {"Low": {"TX": [5, 5], "RX": [5, 5]}, "Mid": {"TX": [3, 7], "RX": [13, 9]}, "Hig": {"TX": [6, 0], "RX": [14, 15]}},
    }
    return fe_list.get(pcb, None)

# Main function that handle read to csv and the right to the excel
def run(pathCSV, pathExcelCopy, pathExcelOuput, temp_list, freq_list_single, mod_bw_offset):

    excel = openpyxl.load_workbook(pathExcelCopy)       # Read excel file
    # Column list given by temperature
    #default columns in excel template --> temp_col = [[7, 10], [13, 16], [19, 22]]
    if temp_list == [55]:
     temp_col = [[13, 16]]
    
    elif temp_list == [-20]:
     temp_col =  [[7, 10]]
    
    elif temp_list == [85]:
     temp_col = [[19, 22]]
    
    elif temp_list == [-20, 55]:
     temp_col = [[7, 10], [13, 16]]
    
    elif temp_list == [-20, 85]:
     temp_col = [[7, 10], [19, 22]] 
    
    elif temp_list == [55, 85]:
     temp_col = [[13, 16], [19, 22]]
    
    elif temp_list == [-20, 55, 85]:
     temp_col = [[7, 10], [13, 16], [19, 22]]
    sheet_list = ['Low_Gain', 'Mid_Gain', 'High_Gain']  # Sheet List
    # Polarization given list
    pol_list = ['V', 'H']
    # Handler for Ref_csv sheet
    csv_sheet = excel['Ref_csv']
    # Get last row from with value or text fromm sheet
    max_row = csv_sheet.max_row

    # Get username name and date then print username
    date_now = datetime.datetime.now()
    username = getpass.getuser()
    print("Signum: " + username.upper())

    # Loop that read one csv file at time in a define folder
    for filename in glob.glob(os.path.join(pathCSV, '*.csv')):
        # Read to a csv fill to a csv handler
        csv = pd.read_csv(filename, sep=',', na_values=['NaN', ' ', '  '])
        print(filename.split('\\')[-1])     # Print csv file name
        pcb_id = csv['PCB_id'][0]           # Get PCB id from csv
        pcb_row = pcb_row_offset(pcb_id)    # Get the row offset
        mode = csv['Mode'][0]               # Get the mode (TX/RX)

        # Get Matrix ids from csv and store in a list
        matrix_id_list = csv['Matrix id'][0]
        matrix_id_list = matrix_id_list.replace(" ", "").split(';')

        # Add data to cvs sheet handler
        max_row += 1
        csv_sheet.cell(
            row=max_row, column=1).value = date_now.strftime("%Y-%m-%d")
        csv_sheet.cell(
            row=max_row, column=2).value = str(matrix_id_list)
        csv_sheet.cell(
            row=max_row, column=3).value = username.upper()
        csv_sheet.cell(
            row=max_row, column=4).value = filename.split('\\')[-1]

        print(matrix_id_list)  # Print Matrix id list

        # Loop excel sheets
        for sheet_name in sheet_list:
            sheet = excel[sheet_name]  # Sheet handler
            temp_pos = 0
            # Loop temapature
            for temp in temp_list:
                # Loop polarization
                for pol in pol_list:
                    # Get the fe list and polarization to the excel
                    if pol == 'V':
                        fe = get_fe_list(pcb_id)[sheet_name[0:3]][mode][0]
                        pol_pos = 0
                    else:
                        fe = get_fe_list(pcb_id)[sheet_name[0:3]][mode][1]
                        pol_pos = 1
                    # Loop matrix id
                    for matrix_id in matrix_id_list:
                        freq_step = 0
                        # Get the frequency list
                        if matrix_id in ['M1', 'M6', 'M7']:
                            freq_list = [freq_list_single[0] + mod_bw_offset,
                                         freq_list_single[1], freq_list_single[2] - mod_bw_offset]
                        else:
                            freq_list = freq_list_single
                        # Loop frequency
                        for freq in freq_list:
                            data = 'NaN'    # Data that will be stored in excel
                            # Filter the csv data
                            temp_data = csv.loc[(csv['Polarization'] == pol) &
                                                (csv['RF Freq'] == freq) & (csv['Temp target'] == temp) & (csv['FE'] == fe)].fillna(100000.00)
                            # Get row where to set the data in excel
                            row = matrix_id_row(matrix_id) + pcb_row

                            # DRS: TX Pout per Branch
                            if matrix_id == 'M1':
                                if temp_data['P_out_ACLR'].any():
                                    data = min(temp_data['P_out_ACLR'])

                            # DRS: TX Gain
                            elif matrix_id == 'M2':
                                data = temp_data['Gain'].values[0]

                            # DRS: TX input CP1dB (ICP1dB)
                            elif matrix_id == 'M3':
                                data = min(temp_data['CPin'])

                            # DRS: TX output CP1dB (OCP1dB)
                            elif matrix_id == 'M4':
                                data = min(temp_data['CPout'])

                            # DRS: TX PSAT
                            elif matrix_id == 'M5':
                                data = min(temp_data['Psat'])

                            # DRS: TX ACLR
                            elif matrix_id == 'M6':
                                if temp_data['ACLR_1_avg'].any():
                                    data = min(temp_data['ACLR_1_avg'])

                            # DRS: TX EVM
                            elif matrix_id == 'M7':
                                temp_data2 = temp_data.loc[(
                                    csv['EVM'] <= 8.0)]
                                p_out = temp_data2['P_out']

                            # DRS: Max TX IF level at Kvint input pin
                            elif matrix_id == 'M8':
                                data = None
                                continue

                            # DRS: TX Off power (Switched off in-band noise)
                            elif matrix_id == 'M9':
                                data = None
                                continue

                            # DRS: TX 4xRFLO signal leakage
                            elif matrix_id == 'M10':
                                continue

                            # DRS: TX 5xRFLO signal leakage
                            elif matrix_id == 'M11':
                                continue

                            # DRS: TX 6xRFLO signal leakage
                            # No 6*LO Cancellation Algorithm
                            elif matrix_id == 'M12':
                                continue

                            # DRS: TX 6xRFLO signal leakage
                            # With 6*LO Cancellation Algorithm
                            elif matrix_id == 'M13':
                                continue

                            # DRS: Diverse TX Spurious Emission
                            elif matrix_id == 'M14':
                                data = None
                                continue

                            # DRS: RX Gain @ 0dB AGC
                            elif matrix_id == 'M15':
                                data = temp_data['Gain'].values[0]

                            # DRS: RX Noise Figure
                            elif matrix_id == 'M16':
                                data = temp_data['NF_fe'].values[0]

                            # DRS: RX in-band output CP1dB (OCP1dB) for max gain
                            elif matrix_id == 'M17':
                                data = min(temp_data['CPout'])

                            # DRS: Diverse RX Spurious Emission & Response Test
                            elif matrix_id == 'M18':
                                data = None
                                continue

                            # In correct matrix id
                            else:
                                raise ValueError(
                                    'matrix_id defined correct')

                            # Convert back empty values
                            if data == 100000.00:
                                data = 'NaN'

                            # Store the result in validation matrix
                            if data is not None:
                                sheet.cell(
                                    row=row, column=temp_col[temp_pos][pol_pos] + freq_step).value = data

                            freq_step += 1  # Incmerment freq step
                temp_pos += 1   # Incmerment temp pos

    sheet = excel['Revision']   # Sheet handler for revision
    # Get last row from with value or text fromm sheet
    max_row = sheet.max_row
    # Get lastest added revission and incoment by 1
    revision = sheet.cell(row=max_row, column=3).value
    revision = int(revision.lstrip('PA'))
    revision += 1

    # Add data to revision sheet handler
    sheet.cell(
        row=max_row + 1, column=2).value = date_now.strftime("%Y-%m-%d")
    sheet.cell(
        row=max_row + 1, column=3).value = "PA" + str(revision)
    sheet.cell(
        row=max_row + 1, column=4).value = username.upper()
    sheet.cell(
        row=max_row + 1, column=5).value = "Validation matrix converter added some data to the Validation matrix"

    # Save the change that all sheet hadler have done to excel file via define path and file name
    excel.save(pathExcelOuput)
